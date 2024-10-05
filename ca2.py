# import openpyxl
# from openpyxl import load_workbook, Workbook
# import hashlib
# import requests
# import logging
# import os
# from ipaddress import ip_address, ip_network

# # Set up logging
# logging.basicConfig(level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')

# # Function to validate email input
# def validate_email_input(email):
#     return '@' in email and '.' in email

# # Function to validate password
# def validate_password(password):
#     if len(password) < 6:
#         print("Password must be at least 6 characters long.")
#         return False
#     return True

# # Function to hash passwords
# def hash_password(password):
#     return hashlib.sha256(password.encode()).hexdigest()

# # Function to read users from the Excel file
# def read_users():
#     try:
#         wb = load_workbook('users.xlsx')
#         ws = wb.active
#         users = []
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             users.append({
#                 'name': row[0],
#                 'email': row[1],
#                 'password': row[2],
#                 'security_question': row[3],
#                 'answer': row[4]
#             })
#         wb.close()
#         return users
#     except Exception as e:
#         logging.error(f"Error reading users: {e}")
#         return []

# # Function to check if an email is already registered
# def is_email_registered(email):
#     users = read_users()
#     for user in users:
#         if user['email'] == email:
#             return True
#     return False

# # Function to register a user
# def register():
#     while True:
#         name = input("Enter your name (or type 'exit' to stop): ").strip()
#         if name.lower() == 'exit':
#             break

#         email = input("Enter your email: ").strip()
#         if not validate_email_input(email):
#             print("Invalid email format. Please try again.")
#             continue

#         # Check if email is already registered
#         if is_email_registered(email):
#             print("This email is already registered. Please try a different email.")
#             continue

#         password = input("Enter your password: ").strip()
#         if not validate_password(password):
#             continue

#         security_question = "What is your favorite place?"
#         answer = input(security_question + ": ").strip()

#         hashed_password = hash_password(password)

#         try:
#             wb = load_workbook('users.xlsx')
#             ws = wb.active
#             ws.append([name, email, hashed_password, security_question, answer])
#             wb.save('users.xlsx')
#             wb.close()  # Ensure file is closed
#             print("User registered successfully!")
#         except PermissionError:
#             logging.error("Permission denied when trying to write to 'users.xlsx'. Check if the file is open elsewhere.")
#             print("Unable to register the user. Please check file permissions and make sure the file is not open in another program.")
#             break
#         except Exception as e:
#             logging.error(f"Unexpected error: {e}")
#             print("An unexpected error occurred. Please try again later.")
#             break

# # Function to login a user with attempt limitation
# def login():
#     max_attempts = 5
#     attempts = 0

#     while attempts < max_attempts:
#         email = input("Enter your email: ").strip()
#         password = input("Enter your password: ").strip()
#         hashed_password = hash_password(password)

#         users = read_users()
#         for user in users:
#             if user['email'] == email and user['password'] == hashed_password:
#                 print("Login successful!")
#                 return user
        
#         attempts += 1
#         remaining_attempts = max_attempts - attempts
#         print(f"Incorrect credentials. {remaining_attempts} attempt(s) remaining.")
        
#         if remaining_attempts == 0:
#             print("Too many failed login attempts. You are locked out of the application.")
#             exit()  # Exit the application after too many attempts

#     return None

# # Function to implement "Forgot Password" with password confirmation
# def forgot_password():
#     email = input("Enter your email: ").strip()
#     users = read_users()

#     for user in users:
#         if user['email'] == email:
#             answer = input(f"{user['security_question']} ").strip()
#             if answer == user['answer']:
#                 while True:
#                     new_password = input("Enter your new password: ").strip()
#                     confirm_password = input("Confirm your new password: ").strip()
                    
#                     if new_password != confirm_password:
#                         print("Passwords do not match. Please try again.")
#                         continue
                    
#                     if validate_password(new_password):
#                         user['password'] = hash_password(new_password)
#                         # Update the users.xlsx with the new password
#                         try:
#                             wb = load_workbook('users.xlsx')
#                             ws = wb.active
#                             for row in ws.iter_rows(min_row=2):
#                                 if row[1].value == email:
#                                     row[2].value = user['password']  # Update the hashed password
#                                     break
#                             wb.save('users.xlsx')
#                             wb.close()
#                             print("Password reset successful!")
#                         except Exception as e:
#                             logging.error(f"Error saving new password: {e}")
#                         return
#                     else:
#                         print("Password does not meet criteria.")
#                         return

#     print("Email or answer to security question is incorrect.")

# # Function to create the users.xlsx file automatically if it doesn't exist
# def create_users_file_if_not_exists():
#     file_path = 'users.xlsx'
#     if not os.path.exists(file_path):
#         wb = Workbook()
#         ws = wb.active
#         # Set headers for the file
#         ws.append(['Name', 'Email', 'Password', 'Security Question', 'Answer'])
#         wb.save(file_path)
#         wb.close()
#         print(f"Created new file: {file_path}")

# # Function to check if IP is private
# def is_private_ip(ip):
#     private_networks = [
#         ip_network('10.0.0.0/8'),
#         ip_network('172.16.0.0/12'),
#         ip_network('192.168.0.0/16')
#     ]
#     ip_addr = ip_address(ip)
#     return any(ip_addr in network for network in private_networks)

# # Function to get geolocation data
# def get_geolocation(ip):
#     if is_private_ip(ip):
#         return {"error": "Private IP detected. No geolocation available."}
    
#     try:
#         response = requests.get(f"https://ipapi.co/{ip}/json/")
#         return response.json()
#     except Exception as e:
#         logging.error(f"Error fetching geolocation: {e}")
#         return {"error": "Unable to retrieve geolocation data."}

# # Function to print geolocation data
# def print_geolocation_data(data):
#     if data.get('error'):
#         print(data['error'])
#     else:
#         print(f"Country: {data.get('country_name')}")
#         print(f"City: {data.get('city')}")
#         print(f"Region: {data.get('region')}")
#         print(f"Latitude: {data.get('latitude')}")
#         print(f"Longitude: {data.get('longitude')}")
#         print(f"Timezone: {data.get('timezone')}")
#         print(f"ISP: {data.get('org')}")

# # Main function to run the application
# def main():
#     # Ensure the users file exists before starting
#     create_users_file_if_not_exists()

#     print("Welcome to IP Geolocation App")
#     while True:
#         print("1. Register")
#         print("2. Login")
#         print("3. Forgot Password")
#         print("4. Exit")
#         choice = input("Enter your choice: ").strip()

#         if choice == '1':
#             register()
#         elif choice == '2':
#             user = login()
#             if user:
#                 ip = input("Enter an IP address (or press enter to use your own): ").strip()
#                 if not ip:
#                     ip = requests.get('https://api64.ipify.org?format=json').json()['ip']
#                 geolocation_data = get_geolocation(ip)
#                 print_geolocation_data(geolocation_data)
#         elif choice == '3':
#             forgot_password()
#         elif choice == '4':
#             break
#         else:
#             print("Invalid choice. Please try again.")

# if __name__ == "__main__":
#     main()


import openpyxl
from openpyxl import load_workbook, Workbook
import hashlib
import requests
import logging
import os
import csv
from datetime import datetime
from ipaddress import ip_address, ip_network

# Set up logging
logging.basicConfig(level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')

# Function to validate email input
def validate_email_input(email):
    return '@' in email and '.' in email

# Function to validate password
def validate_password(password):
    if len(password) < 6:
        print("Password must be at least 6 characters long.")
        return False
    return True

# Function to hash passwords
def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# Function to read users from the Excel file
def read_users():
    try:
        wb = load_workbook('users.xlsx')
        ws = wb.active
        users = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            users.append({
                'name': row[0],
                'email': row[1],
                'password': row[2],
                'security_question': row[3],
                'answer': row[4]
            })
        wb.close()
        return users
    except Exception as e:
        logging.error(f"Error reading users: {e}")
        return []

# Function to check if an email is already registered
def is_email_registered(email):
    users = read_users()
    for user in users:
        if user['email'] == email:
            return True
    return False

# Function to register a user
def register():
    while True:
        name = input("Enter your name (or type 'exit' to stop): ").strip()
        if name.lower() == 'exit':
            break

        email = input("Enter your email: ").strip()
        if not validate_email_input(email):
            print("Invalid email format. Please try again.")
            continue

        # Check if email is already registered
        if is_email_registered(email):
            print("This email is already registered. Please try a different email.")
            continue

        password = input("Enter your password: ").strip()
        if not validate_password(password):
            continue

        security_question = "What is your favorite place?"
        answer = input(security_question + ": ").strip()

        hashed_password = hash_password(password)

        try:
            wb = load_workbook('users.xlsx')
            ws = wb.active
            ws.append([name, email, hashed_password, security_question, answer])
            wb.save('users.xlsx')
            wb.close()  # Ensure file is closed
            print("User registered successfully!")
        except PermissionError:
            logging.error("Permission denied when trying to write to 'users.xlsx'. Check if the file is open elsewhere.")
            print("Unable to register the user. Please check file permissions and make sure the file is not open in another program.")
            break
        except Exception as e:
            logging.error(f"Unexpected error: {e}")
            print("An unexpected error occurred. Please try again later.")
            break

# Function to log login and logout times in CSV
def log_login_logout(email, action):
    file_exists = os.path.isfile('user_log.csv')
    
    with open('user_log.csv', mode='a', newline='') as file:
        writer = csv.writer(file)
        if not file_exists:
            # Write the header if the file is being created
            writer.writerow(['Email', 'Action', 'Timestamp'])
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        writer.writerow([email, action, timestamp])

# Function to login a user with attempt limitation and log login
def login():
    max_attempts = 5
    attempts = 0

    while attempts < max_attempts:
        email = input("Enter your email: ").strip()
        password = input("Enter your password: ").strip()
        hashed_password = hash_password(password)

        users = read_users()
        for user in users:
            if user['email'] == email and user['password'] == hashed_password:
                print("Login successful!")
                log_login_logout(email, "Login")
                return user
        
        attempts += 1
        remaining_attempts = max_attempts - attempts
        print(f"Incorrect credentials. {remaining_attempts} attempt(s) remaining.")
        
        if remaining_attempts == 0:
            print("Too many failed login attempts. You are locked out of the application.")
            exit()  # Exit the application after too many attempts

    return None

# Function to implement logout and log logout
def logout(user):
    log_login_logout(user['email'], "Logout")
    print("You have successfully logged out.")

# Function to implement "Forgot Password" with password confirmation
def forgot_password():
    email = input("Enter your email: ").strip()
    users = read_users()

    for user in users:
        if user['email'] == email:
            answer = input(f"{user['security_question']} ").strip()
            if answer == user['answer']:
                while True:
                    new_password = input("Enter your new password: ").strip()
                    confirm_password = input("Confirm your new password: ").strip()
                    
                    if new_password != confirm_password:
                        print("Passwords do not match. Please try again.")
                        continue
                    
                    if validate_password(new_password):
                        user['password'] = hash_password(new_password)
                        # Update the users.xlsx with the new password
                        try:
                            wb = load_workbook('users.xlsx')
                            ws = wb.active
                            for row in ws.iter_rows(min_row=2):
                                if row[1].value == email:
                                    row[2].value = user['password']  # Update the hashed password
                                    break
                            wb.save('users.xlsx')
                            wb.close()
                            print("Password reset successful!")
                        except Exception as e:
                            logging.error(f"Error saving new password: {e}")
                        return
                    else:
                        print("Password does not meet criteria.")
                        return

    print("Email or answer to security question is incorrect.")

# Function to create the users.xlsx file automatically if it doesn't exist
def create_users_file_if_not_exists():
    file_path = 'users.xlsx'
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        # Set headers for the file
        ws.append(['Name', 'Email', 'Password', 'Security Question', 'Answer'])
        wb.save(file_path)
        wb.close()
        print(f"Created new file: {file_path}")

# Function to check if IP is private
def is_private_ip(ip):
    private_networks = [
        ip_network('10.0.0.0/8'),
        ip_network('172.16.0.0/12'),
        ip_network('192.168.0.0/16')
    ]
    ip_addr = ip_address(ip)
    return any(ip_addr in network for network in private_networks)

# Function to get geolocation data
def get_geolocation(ip):
    if is_private_ip(ip):
        return {"error": "Private IP detected. No geolocation available."}
    
    try:
        response = requests.get(f"https://ipapi.co/{ip}/json/")
        return response.json()
    except Exception as e:
        logging.error(f"Error fetching geolocation: {e}")
        return {"error": "Unable to retrieve geolocation data."}

# Function to print geolocation data
def print_geolocation_data(data):
    if data.get('error'):
        print(data['error'])
    else:
        print(f"Country: {data.get('country_name')}")
        print(f"City: {data.get('city')}")
        print(f"Region: {data.get('region')}")
        print(f"Latitude: {data.get('latitude')}")
        print(f"Longitude: {data.get('longitude')}")
        print(f"Timezone: {data.get('timezone')}")
        print(f"ISP: {data.get('org')}")

# Main function to run the application
def main():
    # Ensure the users file exists before starting
    create_users_file_if_not_exists()

    print("Welcome to IP Geolocation App")
    while True:
        print("1. Register")
        print("2. Login")
        print("3. Forgot Password")
        print("4. Exit")
        choice = input("Enter your choice: ").strip()

        if choice == '1':
            register()
        elif choice == '2':
            user = login()
            if user:
                ip = input("Enter an IP address (or press enter to use your own): ").strip()
                if not ip:
                    ip = requests.get('https://api64.ipify.org?format=json').json()['ip']
                geolocation_data = get_geolocation(ip)
                print_geolocation_data(geolocation_data)
                
                # Log out user after operations
                input("Press enter to log out.")
                logout(user)
        elif choice == '3':
            forgot_password()
        elif choice == '4':
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
